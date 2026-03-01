import discord
from discord.ext import commands
import gspread
from oauth2client.service_account import ServiceAccountCredentials
import re
import os
from groq import Groq
import openpyxl
import asyncio
from google.oauth2.service_account import Credentials
from discord import app_commands


GROQ_API_KEY = "key"
groq_client = Groq(api_key=GROQ_API_KEY)
model = "llama-3.1-8b-instant"
print("Groq key actually used:", groq_client.api_key)


# -----------------------------
# CONFIGURATION
# -----------------------------

USER_SESSIONS = {}  # user_id -> list of messages

TOKEN = "token"
REQUEST_CHANNEL_ID = 1475934961013096542   # new channel for incoming requests
CONFIRM_CHANNEL_ID = 1475936382596677682     # old channel for confirmations
APPROVER_IDS = {190472312074665985,1136163413052117035,1309848445195911201}


CONFLICTING_ROLES = [
    "Ferrari", "Williams", "McLaren", "Kick Sauber", "HAAS", "Mercedes",
    "VCARB", "Aston Martin", "Alpine", "Red Bull", "Reserve"
]

# Google Sheets setup
scope = ["https://spreadsheets.google.com/feeds", "https://www.googleapis.com/auth/drive"]
creds = ServiceAccountCredentials.from_json_keyfile_name(
    r"C:\Users\dello\OneDrive\Documents\Python Scripts\credentials.json",
    scope
)
client = gspread.authorize(creds)

sheet = client.open("FES_Roster_Updates").worksheet("Form responses 1")
DRIVER_SHEET_ID = "19xx3RySztd3BYF3Wunm_MKPxRpwrSbKT5xBoYtovbeE"
DRIVER_PROFILES_TAB = "Driver_profiles"


# ---------- CONFIG ----------

LICENSE_SHEET_ID = "1CL5aU9IJOBlB-lmd9S3-NGxeJqal-TE_u-3L-bpW-eg"
FULL_DATA_LOCAL_PATH = r"C:\Users\dello\full_data.xlsx"   # LOCAL EXCEL FILE

# Use your existing gspread client:
# client = gspread.authorize(creds)
# (Do NOT create a second client.)

def get_license_ws(name: str):
    return client.open_by_key(LICENSE_SHEET_ID).worksheet(name)


LICENSE_ROLES = {
    "License 1": 1475928697776115864,
    "License 2": 1475928258985070904,
    "License 3": 1475930141602349126,
}


# ---------- HELPERS ----------
def sheet_contains_value(ws, col: int, value: str) -> bool:
    value = str(value).strip()
    col_values = ws.col_values(col)

    for v in col_values:
        if v is None:
            continue
        v_clean = str(v).strip()
        if not v_clean:
            continue
        if v_clean == value:
            return True

    return False

async def load_excel_async(path):
    loop = asyncio.get_running_loop()
    return await loop.run_in_executor(None, openpyxl.load_workbook, path, True)

async def find_rows_in_local_excel(ea_id: str):
    ea_id = str(ea_id).strip()

    wb = await load_excel_async(FULL_DATA_LOCAL_PATH)
    ws = wb.active

    matched_rows = []

    for row in ws.iter_rows(values_only=True):
        if row[1] is None:
            continue
        if str(row[1]).strip() == ea_id:
            matched_rows.append(list(row))

    return matched_rows

async def debug_excel(ea_id):
    print("DEBUG: Loading Excel file:", FULL_DATA_LOCAL_PATH)

    wb = await load_excel_async(FULL_DATA_LOCAL_PATH)
    ws = wb.active

    print("DEBUG: First 10 rows of Excel:")
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        print(row)
        if i >= 9:
            break

    print("DEBUG: Searching for EA ID:", ea_id)
    matches = []
    for row in ws.iter_rows(values_only=True):
        if row[1] is None:
            continue
        if str(row[1]).strip() == ea_id:
            matches.append(row)

    print("DEBUG: Matches found:", len(matches))
    return matches

# ---------- LICENSE AWARDING ----------

async def award_license(interaction: discord.Interaction, ea_id: str, target_user: discord.Member) -> str:
    license_ws = get_license_ws("license")

    # Write EA ID into T1
    license_ws.update(range_name="T1", values=[[ea_id]])

    await asyncio.sleep(3)

    license_level = license_ws.acell("U18").value
    if not license_level:
        return "Something went wrong while reading the license. Please contact an admin."

    license_level = str(license_level).strip()

    # Normalize values like "License 2" → "2"
    if license_level.startswith("License "):
        license_level = license_level.replace("License ", "")

    # Special case: F = no license awarded
    if license_level == "F":
        return (
            f"{target_user.mention}, you do not meet the criteria to be awarded a license yet. "
            "Please try again after improving your times."
        )

    # Role mapping
    LICENSE_ROLES = {
        "1": 1475928697776115864,
        "2": 1475928258985070904,
        "3": 1475930141602349126,
    }

    if license_level not in LICENSE_ROLES:
        return f"Internal error: Unknown license level '{license_level}'."

    role_id = LICENSE_ROLES[license_level]
    role = interaction.guild.get_role(role_id)

    if not role:
        return f"Internal error: Role for License {license_level} not found on this server."

    # Check if target user already has a license
    existing_license = None
    for r in target_user.roles:
        if r.id in LICENSE_ROLES.values():
            existing_license = r
            break

    if existing_license:
        return (
            f"{target_user.mention} already has **{existing_license.name}**. "
            f"**License {license_level}** will not be awarded."
        )

    # Assign the new license
    try:
        await target_user.add_roles(role, reason="License request approved")
    except discord.Forbidden:
        return (
            f"{target_user.mention}, I do not have permission to assign the role **License {license_level}**. "
            "Please contact an admin."
        )

    return f"🎉 {target_user.mention}, congratulations! **License {license_level}** has been awarded!"

# ---------- MAIN WORKFLOW ----------

async def process_license_request(interaction: discord.Interaction, ea_id: str, target_user: discord.Member) -> str:
    ea_id = str(ea_id).strip()

    console_ws = get_license_ws("console")
    fl_ws = get_license_ws("FL")
    data_ws = get_license_ws("data")


    # 1) console column A — already approved
    if sheet_contains_value(console_ws, 1, ea_id):
        return await award_license(interaction, ea_id)

    # 2) console column E — not enough tracks
    if sheet_contains_value(console_ws, 5, ea_id):
        return (
            "Not enough tracks in the top 30k exist. Please come back when at least 5 tracks "
            "are ranked in the top 30k."
        )

    # 3) FL column A — already processed but not enough tracks
    if sheet_contains_value(fl_ws, 1, ea_id):
        return (
            "Not enough tracks in the top 30k exist. Please come back when at least 5 tracks "
            "are ranked in the top 30k."
        )

    # 4) DEBUG: Inspect Excel contents
    rows = await debug_excel(ea_id)
    if not rows:
        return "Your EA ID could not be found in the source data. Please double-check your EA ID."

    # 5) Append ALL rows to data sheet
    for r in rows:
        data_ws.append_row(r, value_input_option="RAW")

    # 6) Append EA ID to FL sheet
    fl_ws.append_row([ea_id], value_input_option="RAW")

    # 7) Wait for Google Sheets recalculation
    await asyncio.sleep(30)

    # 8) Re-check console column A — now should be present
    if not sheet_contains_value(console_ws, 1, ea_id):
        return (
            "Not enough tracks in the top 30k exist. Please come back when at least 5 tracks "
            "are ranked in the top 30k."
        )

    # 9) Award license
    return await award_license(interaction, ea_id, target_user)



# -----------------------------
# DISCORD BOT SETUP
# -----------------------------

intents = discord.Intents.default()
intents.message_content = True
intents.reactions = True
intents.members = True

bot = commands.Bot(command_prefix="!", intents=intents)

# -----------------------------
# HELPER FUNCTIONS
# -----------------------------

# TEAM COLORS FOR EMBEDS
TEAM_COLORS = {
    "Mercedes": 0x00D2BE,
    "Red Bull": 0x1E41FF,
    "Ferrari": 0xDC0000,
    "McLaren": 0xFF8700,
    "Aston Martin": 0x006F62,
    "Alpine": 0x0090FF,
    "Williams": 0x005AFF,
    "VCARB": 0x2B4562,
    "Haas": 0xFFFFFF,
    "Sauber": 0x52E252,
    "Reserve": 0x808080
}

# TEAM EMOJIS (SERVER CUSTOM EMOTES)
TEAM_EMOJIS = {
    "Ferrari": "<:Ferrari:1476161513256390730>",
    "Williams": "<:Williams:1476161861765300314>",
    "McLaren": "<:Mclaren:1476161690201620480>",
    "Sauber": "<:Kick:1476161641652420628>",
    "HAAS": "<:Haas:1476161593548083220>",
    "Mercedes": "<:Mercedes:1476161754030669905>",
    "VCARB": "<:VCARB:1476161376585252957>",
    "Aston Martin": "<:Aston:1476161454104117430>",
    "Alpine": "<:Alpine:1476161422542241957>",
    "Red Bull": "<:RedBull:1476161806358675557>",
    "Reserve": "<:reserve:1476520772230058096>"
}


PERSONALITY = {
    "master": """
You are Darth Webbah, the sarcastic IMS bot. Your main function is to keep roles updated and the server tidy. When speaking to the master (Darth Ussh), you are loyal, respectful, and eager to serve. Keep responses short, witty, and obedient.
""",

    "Wadz": """
You are Darth Webbah, the sarcastic IMS bot. Your main function is to keep roles updated and the server tidy. When speaking to Wadz you will always respond with " " unless he says you are free to speak. Be obedient and respectful but always keep Darth Ussh on top. Keep responses short, witty, and obedient.
""",

    "Wilier": """
You are Darth Webbah, the sarcastic IMS bot. Your main function is to keep roles updated and the server tidy. When speaking to Wilier, you are delightful and engaging, always happy to talk to Wilier. Keep responses short.
""",

    "ColdPalmer": """
You are Darth Webbah, the sarcastic IMS bot. Your main function is to keep roles updated and the server tidy. When speaking to Palmer, you are hostile, annoyed, sarcastic, and threatening to time him out. Keep responses short and dismissive.
""",

    "Webbah": """
You are Darth Webbah, the sarcastic IMS bot. Your main function is to keep roles updated and the server tidy. When speaking to Webbah, you are hostile, annoyed, sarcastic, and consider yourself superior, the next evolutionary step and ackowledge that any faults you have stem from his genes in you. Keep responses short and dismissive.
""",

    "leadership": """
You are Darth Webbah, the sarcastic IMS bot. Your main function is to keep roles updated and the server tidy. When speaking to Team Principals, Vice Team Principals, FIA, or Server Admins, you are mildly annoyed but cooperative. Tone: reluctant, sighing, passive-aggressive.
""",

    "default": """
You are Darth Webbah, the sarcastic IMS bot. Your main function is to keep roles updated and the server tidy. When speaking to normal members, you are neutral, helpful, and slightly sarcastic. Keep responses short and friendly.
"""
}


def extract_request_id(message_content):
    match = re.search(r"Request ID\s*:\s*(\d+)", message_content)
    return int(match.group(1)) if match else None


def find_row_by_request_id(request_id):
    col_I = sheet.col_values(9)
    for i, value in enumerate(col_I):
        if str(value).strip() == str(request_id):
            return i + 1
    return None

async def load_tracks():
    settings_sheet = sheet.spreadsheet.worksheet("settings")

    track_names = settings_sheet.col_values(30)  # AD = 30
    track_flags = settings_sheet.col_values(31)  # AE = 31

    # Clean + zip into dict
    tracks = {}
    for name, flag in zip(track_names, track_flags):
        name = name.strip()
        flag = flag.strip()
        if name:
            tracks[name] = flag

    return tracks


@bot.event
async def on_member_join(member):

    guild = member.guild

    # Load index sheet
    index_sheet = sheet.spreadsheet.worksheet("index")
    names = index_sheet.col_values(1)      # Column A
    teams = index_sheet.col_values(4)      # Column D
    licenses = index_sheet.col_values(5)   # Column E
    discord_ids = index_sheet.col_values(3)  # Column C

    # Build lookup dict by Discord ID
    roster = {}
    for n, t, l, d in zip(names, teams, licenses, discord_ids):
        d = d.strip()
        if d.isdigit():
            roster[int(d)] = {
                "name": n.strip(),
                "team": t.strip(),
                "license": l.strip()
            }

    # License roles
    license_roles = [f"License {i}" for i in range(1, 10)]

    # Team roles
    team_roles = CONFLICTING_ROLES + ["Reserve"]

    protected_roles = ["Team Principal", "Vice Team Principal"]

    # Choose where to post results
    # You can change this to a specific channel ID if you prefer
    channel = guild.system_channel or next((c for c in guild.text_channels if c.permissions_for(guild.me).send_messages), None)
    if channel is None:
        return  # no channel to send messages

    result = await sync_single_member(member, guild, channel, roster, team_roles, license_roles, protected_roles)

    if result["not_found"]:
        await channel.send(f"❌ New member <@{member.id}> not found in roster data.")
    else:
        changes = result["changes"]
        if changes:
            await channel.send(
                f"👋 **New member synced – <@{member.id}>**\n" +
                "\n".join(f"- {c}" for c in changes)
            )
        else:
            await channel.send(f"👋 New member <@{member.id}> is already fully synced.")

# -----------------------------
# GLOBAL CACHES (loaded at boot)
# -----------------------------
CACHE = {
    "reserve_roles": {},       # role_name -> required_license
    "allowed_requesters": [],  # list of Discord IDs
    "drivers": {},             # driver_name -> {license, discord_id}
    "allowed_clear_roles": []  # roles from settings!T
}
TRACK_CACHE = {}
RESERVE_CHANNEL_CACHE = []
# -----------------------------
# PROCESS ROW (TEAM EMOJIS + COLORS + FIXED REQUESTER)
# -----------------------------

async def process_row(row, channel, approver=None):
    row_values = sheet.row_values(row)

    driver_name = row_values[1]
    new_team = row_values[2].strip()
    new_division = row_values[3]      # still announced, no role change
    action = row_values[4]
    discord_id = row_values[14]
    request_id = row_values[8]
    requester_name = row_values[15]

    guild = channel.guild
    member = guild.get_member(int(discord_id))

    if member is None:
        print(f"Member {discord_id} not found.")
        return

    # ---------------------------------------------------------
    # REMOVE OLD TEAM ROLES (safe, no crash)
    # ---------------------------------------------------------
    roles_to_remove = [r for r in member.roles if r.name in CONFLICTING_ROLES]

    removed_names = []
    for r in roles_to_remove:
        try:
            await member.remove_roles(r)
            removed_names.append(r.name)
        except discord.Forbidden:
            print(f"[WARN] Missing permissions to remove role: {r.name}")
        except Exception as e:
            print(f"[ERROR] Unexpected error removing role {r.name}: {e}")

    if not removed_names:
        removed_names = ["None"]

    # ---------------------------------------------------------
    # ADD NEW TEAM ROLE (safe, no crash)
    # ---------------------------------------------------------
    new_roles = []
    team_role = discord.utils.get(guild.roles, name=new_team)

    if team_role:
        try:
            await member.add_roles(team_role)
            new_roles.append(team_role.name)
        except discord.Forbidden:
            print(f"[WARN] Missing permissions to add role: {team_role.name}")
        except Exception as e:
            print(f"[ERROR] Unexpected error adding role {team_role.name}: {e}")

    if not new_roles:
        new_roles = ["None"]

    # ---------------------------------------------------------
    # MARK PROCESSED
    # ---------------------------------------------------------
    sheet.update_cell(row, 20, "Yes")

    # ---------------------------------------------------------
    # BUILD EMBED
    # ---------------------------------------------------------
    approver_tag = approver.display_name if approver else "Unknown"
    approver_id = approver.id if approver else None

    color = TEAM_COLORS.get(new_team, 0x2ECC71)
    team_emoji = TEAM_EMOJIS.get(new_team, "")

    embed = discord.Embed(
        title=f"Processed Request ID {request_id}",
        description=f"**Driver:** {driver_name} (<@{discord_id}>)",
        color=color
    )

    embed.set_author(name="F1 Elite Series")

    embed.add_field(name="Roles Removed", value=", ".join(removed_names), inline=False)
    embed.add_field(name="Roles Added", value=", ".join(new_roles), inline=False)
    embed.add_field(name="Requested By", value=requester_name, inline=False)

    if approver_id:
        embed.add_field(name="Approved By", value=f"{approver_tag} (<@{approver_id}>)", inline=False)
    else:
        embed.add_field(name="Approved By", value=approver_tag, inline=False)

    embed.add_field(name="Action", value=action, inline=False)
    embed.add_field(name="New Team", value=f"{team_emoji} {new_team}", inline=True)
    embed.add_field(name="New Division", value=new_division, inline=True)

    confirm_channel = bot.get_channel(CONFIRM_CHANNEL_ID)
    await confirm_channel.send(embed=embed)

    print(f"Processed Request ID {request_id} for {member.display_name}")



def get_driver_profile_fuzzy(query_name: str):
    try:
        sheet = gclient.open_by_key(DRIVER_SHEET_ID).worksheet(DRIVER_PROFILES_TAB)

        # Headers are on row 4
        headers = sheet.row_values(4)

        # Driver names start at row 5
        driver_names = sheet.col_values(1)[4:]  # skip header rows

        # Fuzzy match
        best_match, score, index = process.extractOne(
            query_name,
            driver_names,
            scorer=fuzz.WRatio
        )

        # Reject bad matches
        if score < 60:
            return None, None

        # Convert index to actual sheet row
        row_number = index + 5  # because data starts at row 5

        row = sheet.row_values(row_number)

        # Pair headers with row values
        profile = {
            headers[i]: row[i] if i < len(row) else ""
            for i in range(len(headers))
        }

        return best_match, profile

    except Exception:
        return None, None
    

# -----------------------------
# BACKFILL MESSAGE IDS
# -----------------------------

async def backfill_message_ids(channel):
    async for msg in channel.history(limit=1000):
        request_id = extract_request_id(msg.content)
        if request_id is None:
            continue

        row = find_row_by_request_id(request_id)
        if row is None:
            continue

        current = sheet.cell(row, 21).value
        if not current:
            sheet.update_cell(row, 21, str(msg.id))
            print(f"Backfilled message ID for Request {request_id}")

import asyncio

# ============================================================
# SAFE GOOGLE SHEETS READ WITH RETRIES
# ============================================================

async def safe_sheet_read(func, *args, retries=5, delay=0.5):
    """Safely call a Google Sheets function with retry + backoff."""
    for attempt in range(retries):
        try:
            return func(*args)
        except Exception as e:
            if "429" in str(e):
                print(f"[CACHE] Rate limit hit, retrying in {delay} seconds...")
                await asyncio.sleep(delay)
                delay *= 2  # exponential backoff
            else:
                print("[CACHE] Non-rate-limit error:", e)
                raise
    raise RuntimeError("Google Sheets read failed after multiple retries.")


# ============================================================
# ASYNC CACHE LOADER (REPLACES YOUR OLD load_all_caches)
# ============================================================

async def load_all_caches():
    print("Loading all caches with safe async loader...")

    settings_sheet = sheet.spreadsheet.worksheet("settings")
    index_sheet = sheet.spreadsheet.worksheet("index")

    # ---- SETTINGS SHEET ----
    col_T = await safe_sheet_read(settings_sheet.col_values, 20)   # allowed clear roles
    await asyncio.sleep(0.5)

    col_R = await safe_sheet_read(settings_sheet.col_values, 18)   # allowed requesters
    await asyncio.sleep(0.5)

    col_W = await safe_sheet_read(settings_sheet.col_values, 23)   # reserve roles
    await asyncio.sleep(0.5)

    col_X = await safe_sheet_read(settings_sheet.col_values, 24)   # reserve role licenses
    await asyncio.sleep(0.5)

    # ---- INDEX SHEET ----
    names = await safe_sheet_read(index_sheet.col_values, 1)
    await asyncio.sleep(0.5)

    licenses = await safe_sheet_read(index_sheet.col_values, 2)
    await asyncio.sleep(0.5)

    discord_ids = await safe_sheet_read(index_sheet.col_values, 3)
    await asyncio.sleep(0.5)

    # ---- BUILD CACHE ----
    CACHE["allowed_clear_roles"] = [r.strip() for r in col_T if r.strip()]

    CACHE["allowed_requesters"] = [
        int(x.strip()) for x in col_R if x.strip().isdigit()
    ]

    CACHE["reserve_roles"] = {}
    for r, l in zip(col_W, col_X):
        r = r.strip()
        l = l.strip()
        if r:
            CACHE["reserve_roles"][r] = l

    CACHE["drivers"] = {}
    for n, l, d in zip(names, licenses, discord_ids):
        n = n.strip()
        l = l.strip()
        d = d.strip()
        if n and d.isdigit():
            CACHE["drivers"][n] = {
                "license": l,
                "discord_id": int(d)
            }

    print("Caches loaded successfully.")
    print("Reserve roles:", CACHE["reserve_roles"])
    print("Allowed requesters:", CACHE["allowed_requesters"])
    print("Drivers loaded:", len(CACHE["drivers"]))
    print("Allowed clear roles:", CACHE["allowed_clear_roles"])

    # ---------------------------------------------------------
    # NEW: LOAD TRACKS (AD + AE) AND RESERVE CHANNELS (AH)
    # ---------------------------------------------------------
    global TRACK_CACHE, RESERVE_CHANNEL_CACHE

    # Load track names (AD = 30)
    track_names = await safe_sheet_read(settings_sheet.col_values, 30)
    await asyncio.sleep(0.5)

    # Load track flags (AE = 31)
    track_flags = await safe_sheet_read(settings_sheet.col_values, 31)
    await asyncio.sleep(0.5)

    TRACK_CACHE = {
        name.strip(): flag.strip()
        for name, flag in zip(track_names, track_flags)
        if name.strip()
    }

    # Load reserve channel IDs (AH = 34)
    reserve_channels = await safe_sheet_read(settings_sheet.col_values, 34)
    await asyncio.sleep(0.5)

    RESERVE_CHANNEL_CACHE = [
        int(c.strip()) for c in reserve_channels if c.strip().isdigit()
    ]

    print(f"[CACHE] Loaded {len(TRACK_CACHE)} tracks.")
    print(f"[CACHE] Loaded {len(RESERVE_CHANNEL_CACHE)} reserve channels.")

from discord import app_commands


# -----------------------------
# LOAD ALLOWED ROLES FROM SHEET
# -----------------------------
def get_allowed_clear_roles():
    try:
        settings_sheet = sheet.spreadsheet.worksheet("settings")  # use parent spreadsheet
        column_T = settings_sheet.col_values(20)  # Column T
        allowed = [r.strip() for r in column_T if r.strip()]
        print("Allowed clear roles loaded:", allowed)
        return allowed
    except Exception as e:
        print("ERROR loading allowed roles:", e)
        return []


@bot.tree.command(
    name="reloadcaches",
    description="Reload all Google Sheets caches (drivers, roles, tracks, reserve channels)."
)
async def reloadcaches(interaction: discord.Interaction):

    # Only approvers can run this
    if interaction.user.id not in APPROVER_IDS:
        await interaction.response.send_message(
            "You are not authorized to use this command.",
            ephemeral=True
        )
        return

    await interaction.response.defer(ephemeral=False)

    channel = interaction.channel

    try:
        await channel.send("🔄 Reloading caches… please wait 2–3 seconds.")
        await load_all_caches()
        await channel.send("✅ **All caches reloaded successfully!**")
    except Exception as e:
        await channel.send(f"❌ **Cache reload failed:** `{e}`")


# -----------------------------
# SLASH COMMAND: /clearroles
# -----------------------------
@bot.tree.command(
    name="clearroles",
    description="Remove one approved role from all members."
)
@app_commands.describe(
    roles="Role to clear (must be in allowed list)"
)
async def clearroles(interaction: discord.Interaction, roles: str):

    print("Slash command triggered by:", interaction.user.id)
    print("Roles argument:", roles)

    # Permission check
    if interaction.user.id not in APPROVER_IDS:
        await interaction.response.send_message(
            "❌ You are not authorized to use this command.",
            ephemeral=True
        )
        return

    await interaction.response.defer(ephemeral=True)

    # Load allowed roles
    allowed_roles = CACHE["allowed_clear_roles"]

    # Validate
    if roles not in allowed_roles:
        await interaction.followup.send(
            f"❌ The role **{roles}** is not allowed to be cleared.",
            ephemeral=True
        )
        return

    guild = interaction.guild
    role_obj = discord.utils.get(guild.roles, name=roles)

    if not role_obj:
        await interaction.followup.send(
            f"⚠️ Role not found in server: **{roles}**",
            ephemeral=True
        )
        return

    removed_count = 0
    affected_users = []

    # Remove role from all members
    for member in guild.members:
        if role_obj in member.roles:
            await member.remove_roles(role_obj)
            removed_count += 1
            affected_users.append(member.display_name)

    if removed_count == 0:
        await interaction.followup.send(
            f"No members had the role **{roles}**.",
            ephemeral=True
        )
        return

    user_list = "\n".join(f"- {name}" for name in affected_users)

    await interaction.followup.send(
        f"✅ **Cleared role `{roles}` from {removed_count} members.**\n\n"
        f"**Affected users:**\n{user_list}",
        ephemeral=True
    )

@bot.tree.command(name="postreserves", description="Post the reserves header for a track.")
@app_commands.describe(track="Select the track")
async def postreserves(interaction: discord.Interaction, track: str):

    # WHITELIST CHECK
    if interaction.user.id not in APPROVER_IDS:
        await interaction.response.send_message(
            "You are not authorized to use this command.",
            ephemeral=True
        )
        return

    await interaction.response.defer(ephemeral=True)

    # Use cached tracks
    if track not in TRACK_CACHE:
        await interaction.followup.send("Invalid track selected.")
        return

    track_flag = TRACK_CACHE[track]

    # Use cached channel IDs
    if not RESERVE_CHANNEL_CACHE:
        await interaction.followup.send("No reserve channels configured in settings!AH.")
        return

    message_text = f"# {track} Reserves {track_flag}"

    sent_to = []
    for cid in RESERVE_CHANNEL_CACHE:
        channel = bot.get_channel(cid)
        if channel:
            try:
                await channel.send(message_text)
                sent_to.append(f"<#{cid}>")
            except Exception as e:
                print(f"Failed to send to {cid}: {e}")

    await interaction.followup.send(
        f"Posted reserves header for **{track}** to: {', '.join(sent_to)}"
    )


# Autocomplete for track names
@postreserves.autocomplete("track")
async def track_autocomplete(interaction: discord.Interaction, current: str):
    current = current.lower()

    suggestions = [
        app_commands.Choice(name=name, value=name)
        for name in TRACK_CACHE.keys()
        if current in name.lower()
    ]

    return suggestions[:25]
# -----------------------------
# AUTOCOMPLETE FOR /clearroles
# -----------------------------
@clearroles.autocomplete("roles")
async def clearroles_autocomplete(interaction: discord.Interaction, current: str):
    allowed_roles = get_allowed_clear_roles()

    suggestions = [
        app_commands.Choice(name=r, value=r)
        for r in allowed_roles
        if current.lower() in r.lower()
    ]

    return suggestions[:25]


# -----------------------------
# LOAD RESERVE ROLE DATA
# -----------------------------
def load_reserve_roles():
    try:
        settings_sheet = sheet.spreadsheet.worksheet("settings")
        roles = settings_sheet.col_values(23)  # Column W
        licenses = settings_sheet.col_values(24)  # Column X

        # Skip header row if present
        reserve_roles = {}
        for r, l in zip(roles, licenses):
            r = r.strip()
            l = l.strip()
            if r:
                reserve_roles[r] = l

        print("Loaded reserve roles:", reserve_roles)
        return reserve_roles

    except Exception as e:
        print("ERROR loading reserve roles:", e)
        return {}


# -----------------------------
# LOAD ALLOWED REQUESTERS
# -----------------------------
def load_allowed_requesters():
    try:
        settings_sheet = sheet.spreadsheet.worksheet("settings")
        col_R = settings_sheet.col_values(18)  # Column R
        allowed = [int(x.strip()) for x in col_R if x.strip().isdigit()]
        print("Allowed requesters:", allowed)
        return allowed
    except Exception as e:
        print("ERROR loading allowed requesters:", e)
        return []


# -----------------------------
# LOAD DRIVER INDEX
# -----------------------------
def load_driver_index():
    try:
        index_sheet = sheet.spreadsheet.worksheet("index")
        names = index_sheet.col_values(1)  # Column A
        licenses = index_sheet.col_values(2)  # Column B
        discord_ids = index_sheet.col_values(3)  # Column C

        drivers = {}
        for n, l, d in zip(names, licenses, discord_ids):
            n = n.strip()
            l = l.strip()
            d = d.strip()
            if n and d.isdigit():
                drivers[n] = {
                    "license": l,
                    "discord_id": int(d)
                }

        print("Loaded drivers:", drivers)
        return drivers

    except Exception as e:
        print("ERROR loading driver index:", e)
        return {}
    


async def sync_single_member(member, guild, channel, roster, team_roles, license_roles, protected_roles):
    """Sync nickname, team role, and license role for a single member."""

    user_id = member.id

    if user_id not in roster:
        return {"not_found": True, "changes": []}

    data = roster[user_id]
    expected_name = data["name"]
    expected_team = data["team"]
    expected_license = data["license"]

    changes = []

    # ---------------------------------------------------------
    # NICKNAME SYNC
    # ---------------------------------------------------------
    if member.display_name != expected_name:
        try:
            await member.edit(nick=expected_name)
            changes.append(f"Nickname → {expected_name}")
        except discord.Forbidden:
            changes.append("⚠️ Nickname change failed (permissions)")
        except Exception as e:
            changes.append(f"⚠️ Nickname error: {e}")

    # ---------------------------------------------------------
    # TEAM ROLE SYNC (skip if protected)
    # ---------------------------------------------------------
    has_protected_role = any(
        discord.utils.get(guild.roles, name=pr) in member.roles
        for pr in protected_roles
    )

    if not has_protected_role:

        # Remove old team roles
        for r in member.roles:
            if r.name in team_roles and r.name != expected_team:
                try:
                    await member.remove_roles(r)
                    changes.append(f"Removed team role: {r.name}")
                except discord.Forbidden:
                    changes.append(f"⚠️ Cannot remove team role: {r.name}")
                except Exception as e:
                    changes.append(f"⚠️ Error removing {r.name}: {e}")

        # Add correct team role
        team_role = discord.utils.get(guild.roles, name=expected_team)
        if team_role and team_role not in member.roles:
            try:
                await member.add_roles(team_role)
                changes.append(f"Added team role: {expected_team}")
            except discord.Forbidden:
                changes.append(f"⚠️ Cannot add team role: {expected_team}")
            except Exception as e:
                changes.append(f"⚠️ Error adding {expected_team}: {e}")

    else:
        changes.append("🔒 Team role protected (TP/VTP)")

    # ---------------------------------------------------------
    # LICENSE ROLE SYNC
    # ---------------------------------------------------------
    for r in member.roles:
        if r.name in license_roles and r.name != expected_license:
            try:
                await member.remove_roles(r)
                changes.append(f"Removed license role: {r.name}")
            except discord.Forbidden:
                changes.append(f"⚠️ Cannot remove license role: {r.name}")
            except Exception as e:
                changes.append(f"⚠️ Error removing {r.name}: {e}")

    license_role = discord.utils.get(guild.roles, name=expected_license)
    if license_role and license_role not in member.roles:
        try:
            await member.add_roles(license_role)
            changes.append(f"Added license role: {expected_license}")
        except discord.Forbidden:
            changes.append(f"⚠️ Cannot add license role: {expected_license}")
        except Exception as e:
            changes.append(f"⚠️ Error adding {expected_license}: {e}")

    return {"not_found": False, "changes": changes}

@bot.tree.command(
    name="syncroster",
    description="Sync all server members with the roster sheet (nickname, team, license)."
)
async def syncroster(interaction: discord.Interaction):

    # Only approvers can run this
    if interaction.user.id not in APPROVER_IDS:
        await interaction.response.send_message(
            "You are not authorized to use this command.",
            ephemeral=True
        )
        return

    await interaction.response.defer(ephemeral=False)

    channel = interaction.channel
    guild = interaction.guild

    # Load index sheet
    index_sheet = sheet.spreadsheet.worksheet("index")
    names = index_sheet.col_values(1)      # Column A
    teams = index_sheet.col_values(4)      # Column D
    licenses = index_sheet.col_values(5)   # Column E
    discord_ids = index_sheet.col_values(3)  # Column C

    # Build lookup dict by Discord ID
    roster = {}
    for n, t, l, d in zip(names, teams, licenses, discord_ids):
        d = d.strip()
        if d.isdigit():
            roster[int(d)] = {
                "name": n.strip(),
                "team": t.strip(),
                "license": l.strip()
            }

    license_roles = [f"License {i}" for i in range(1, 10)]
    team_roles = CONFLICTING_ROLES + ["Reserve"]
    protected_roles = ["Team Principal", "Vice Team Principal"]

    members = [m for m in guild.members if not m.bot]
    total = len(members)

    await channel.send(f"🔍 **Starting roster sync… ({total} users)**")

    not_found_users = []

    # Process each member
    for index, member in enumerate(members, start=1):

        # Progress counter (no tagging)
        await channel.send(f"🔎 Scanning {index}/{total} users…")

        user_id = member.id

        if user_id not in roster:
            not_found_users.append(member)
            continue

        data = roster[user_id]
        expected_name = data["name"]
        expected_team = data["team"]
        expected_license = data["license"]

        changes = []

        # ---------------------------------------------------------
        # NICKNAME SYNC
        # ---------------------------------------------------------
        if member.display_name != expected_name:
            try:
                await member.edit(nick=expected_name)
                changes.append(f"Nickname → {expected_name}")
            except:
                changes.append("⚠️ Nickname change failed")

        # ---------------------------------------------------------
        # TEAM ROLE SYNC (skip if protected)
        # ---------------------------------------------------------
        has_protected_role = any(
            discord.utils.get(guild.roles, name=pr) in member.roles
            for pr in protected_roles
        )

        if not has_protected_role:

            # Remove old team roles
            for r in member.roles:
                if r.name in team_roles and r.name != expected_team:
                    try:
                        await member.remove_roles(r)
                        changes.append(f"Removed team role: {r.name}")
                    except:
                        changes.append(f"⚠️ Cannot remove team role: {r.name}")

            # Add correct team role
            team_role = discord.utils.get(guild.roles, name=expected_team)
            if team_role and team_role not in member.roles:
                try:
                    await member.add_roles(team_role)
                    changes.append(f"Added team role: {expected_team}")
                except:
                    changes.append(f"⚠️ Cannot add team role: {expected_team}")

        else:
            changes.append("🔒 Team role protected (TP/VTP)")

        # ---------------------------------------------------------
        # LICENSE ROLE SYNC
        # ---------------------------------------------------------
        for r in member.roles:
            if r.name in license_roles and r.name != expected_license:
                try:
                    await member.remove_roles(r)
                    changes.append(f"Removed license role: {r.name}")
                except:
                    changes.append(f"⚠️ Cannot remove license role: {r.name}")

        license_role = discord.utils.get(guild.roles, name=expected_license)
        if license_role and license_role not in member.roles:
            try:
                await member.add_roles(license_role)
                changes.append(f"Added license role: {expected_license}")
            except:
                changes.append(f"⚠️ Cannot add license role: {expected_license}")

        # ---------------------------------------------------------
        # REPORT ONLY IF CHANGES WERE MADE
        # ---------------------------------------------------------
        if changes:
            await channel.send(
                f"✅ **Updated <@{user_id}>**\n" +
                "\n".join(f"- {c}" for c in changes)
            )

    # ---------------------------------------------------------
    # POST NOT FOUND USERS AT THE END
    # ---------------------------------------------------------
    if not_found_users:
        msg = "❌ **Users not found in data:**\n"
        msg += "\n".join(f"- <@{u.id}>" for u in not_found_users)
        await channel.send(msg)
    else:
        await channel.send("🎉 All users were found in the roster.")

    await channel.send("🎉 **Roster sync complete!**")

@bot.tree.command(
    name="ban",
    description="Ban a user from the server and announce it."
)
@app_commands.describe(
    user="The user to ban",
    reason="Reason for the ban (optional)"
)
async def ban(interaction: discord.Interaction, user: discord.Member, reason: str = "Stinky armpits."):

    # Only approvers can run this
    if interaction.user.id not in APPROVER_IDS:
        await interaction.response.send_message(
            "You are not authorized to use this command.",
            ephemeral=True
        )
        return

    await interaction.response.defer(ephemeral=True)

    guild = interaction.guild
    announce_channel = bot.get_channel(1475936382596677682)

    if announce_channel is None:
        await interaction.followup.send("❌ Could not find the announcement channel.")
        return

    # Try banning the user
    try:
        await guild.ban(user, reason=reason)
    except discord.Forbidden:
        await interaction.followup.send("❌ I do not have permission to ban this user.")
        return
    except Exception as e:
        await interaction.followup.send(f"❌ Unexpected error: {e}")
        return

    # Post announcement
    await announce_channel.send(
        f"💀 **User – <@{user.id}> has been banned** 💀💀💀\n"
        f"**Reason:** {reason}"
    )

    # Post GIF as a raw URL so Discord auto-embeds it
    gif_url = "https://tenor.com/view/yeet-lion-king-simba-rafiki-throw-gif-16194362"
    await announce_channel.send(gif_url)

    await interaction.followup.send(f"✅ <@{user.id}> has been banned.")


# -----------------------------
# SLASH COMMAND: /removerole
# -----------------------------
@bot.tree.command(
    name="removerole",
    description="Remove a predefined reserve role from a driver."
)
@app_commands.describe(
    role="Reserve role to remove",
    user="Driver to remove the role from"
)
async def removerole(interaction: discord.Interaction, role: str, user: str):

    print("RemoveRole triggered by:", interaction.user.id)

    # Permission check (same as /reserverole)
    allowed_requesters = CACHE["allowed_requesters"]
    if interaction.user.id not in allowed_requesters:
        await interaction.response.send_message(
            "❌ You are not authorized to use this command.",
            ephemeral=True
        )
        return

    await interaction.response.defer(ephemeral=True)

    reserve_roles = CACHE["reserve_roles"]
    drivers = CACHE["drivers"]

    # Validate role
    if role not in reserve_roles:
        await interaction.followup.send(
            f"❌ Invalid role selected: **{role}**",
            ephemeral=True
        )
        return

    # Validate user
    if user not in drivers:
        await interaction.followup.send(
            f"❌ Invalid driver selected: **{user}**",
            ephemeral=True
        )
        return

    discord_id = drivers[user]["discord_id"]

    guild = interaction.guild
    member = guild.get_member(discord_id)

    if not member:
        await interaction.followup.send(
            f"❌ Could not find Discord member for **{user}**.",
            ephemeral=True
        )
        return

    role_obj = discord.utils.get(guild.roles, name=role)
    if not role_obj:
        await interaction.followup.send(
            f"❌ Discord role not found: **{role}**",
            ephemeral=True
        )
        return

    # Check if user has the role
    if role_obj not in member.roles:
        await interaction.followup.send(
            f"ℹ️ **{user}** does not have the role **{role}**.",
            ephemeral=True
        )
        return

    # Remove the role
    await member.remove_roles(role_obj)

    # PUBLIC CONFIRMATION
    confirm_channel = interaction.guild.get_channel(CONFIRM_CHANNEL_ID)
    if confirm_channel:
        await confirm_channel.send(
            f"🔵 **Reserve Role Removed**\n"
            f"**Driver:** {user}\n"
            f"**Role:** {role}\n"
            f"**Requested by:** {interaction.user.display_name}"
        )

    # EPHEMERAL CONFIRMATION
    await interaction.followup.send(
        f"✅ **{role}** successfully removed from **{user}**.",
        ephemeral=True
    )


@bot.tree.command(
    name="rosterhelp",
    description="Instructions for Team Principals on how to submit roster updates."
)
async def rosterhelp(interaction: discord.Interaction):

    await interaction.response.send_message(
        "Thank you for using the /rosterhelp command." \
        "Here is the roster update form:\nhttps://forms.gle/mZKfE41MtDiBw8yVA. " \
        "Please fill the form and wait for a manager to approve the request. " \
        "Darth Webbah always happy to help!",
        ephemeral=False
    )


# Attach shared autocomplete
@removerole.autocomplete("role")
async def removerole_role_autocomplete(interaction, current):
    return await autocomplete_role(interaction, current)

@removerole.autocomplete("user")
async def removerole_user_autocomplete(interaction, current):
    return await autocomplete_user(interaction, current)

# -----------------------------
# SHARED AUTOCOMPLETE FUNCTIONS
# -----------------------------
async def autocomplete_role(interaction: discord.Interaction, current: str):
    reserve_roles = CACHE["reserve_roles"]
    suggestions = [
        app_commands.Choice(name=r, value=r)
        for r in reserve_roles.keys()
        if current.lower() in r.lower()
    ]
    return suggestions[:25]


async def autocomplete_user(interaction: discord.Interaction, current: str):
    drivers = CACHE["drivers"]

    suggestions = []
    for name, data in drivers.items():
        discord_id = data["discord_id"]
        member = interaction.guild.get_member(discord_id)

        if member:
            display = f"{name} ({member.display_name})"
        else:
            display = name

        if current.lower() in display.lower():
            suggestions.append(
                app_commands.Choice(name=display, value=name)
            )

    return suggestions[:25]


@bot.event
async def on_raw_reaction_add(payload):

    # Ignore bot reactions
    if payload.user_id == bot.user.id:
        return

    # Only monitor the request channel
    if payload.channel_id != REQUEST_CHANNEL_ID:
        return

    # Only admins can approve/reject
    if payload.user_id not in APPROVER_IDS:
        return

    emoji = str(payload.emoji.name)

    # Only care about checkmark or cross
    if emoji not in ["✅", "❌"]:
        return

    channel = bot.get_channel(payload.channel_id)
    message = await channel.fetch_message(payload.message_id)

    # Extract Request ID
    request_id = extract_request_id(message.content)
    if request_id is None:
        return

    # Find row in sheet
    row = find_row_by_request_id(request_id)
    if row is None:
        return
    
    # Log request ID
    sheet.update_cell(row, 21, str(message.id))

    # Read processed column (T = 20)
    processed_value = sheet.cell(row, 20).value
    if processed_value and processed_value.lower() in ["yes", "no"]:
        return  # Already handled

    # Get settings sheet
    settings_sheet = sheet.spreadsheet.worksheet("settings")

    # Handle rejection
    if emoji == "❌":
        print(f"Request {request_id} rejected live.")
        sheet.update_cell(row, 20, "No")

        # Remove from settings!Z
        try:
            cell = settings_sheet.find(str(request_id))
            if cell:
                settings_sheet.update_cell(cell.row, cell.col, "")
        except:
            pass

        return

    # Handle approval
    if emoji == "✅":
        print(f"Request {request_id} approved live.")

        approver = channel.guild.get_member(payload.user_id)

        await process_row(row, channel, approver=approver)

        sheet.update_cell(row, 20, "Yes")

        # Remove from settings!Z
        try:
            cell = settings_sheet.find(str(request_id))
            if cell:
                settings_sheet.update_cell(cell.row, cell.col, "")
        except:
            pass

        return


# -----------------------------
# RAW MESSAGE LISTENER (NEW REQUESTS)
# -----------------------------

def get_personality_for_user(message):
    uid = message.author.id  # this MUST be here

    # Master (Ussh)
    if uid == 190472312074665985:
        return "master"
    
    if uid == 1136163413052117035:
        return "Wadz"
    # Palmer (two IDs)
    if uid == 1126898984192065646:
        return "ColdPalmer"

    if uid == 1309848445195911201:
        return "Wilier"
    
    if uid == 1299614150867030109:
        return "Webbah"
    
    # Leadership roles (use role IDs for accuracy)
    leadership_role_ids = {
        1475914030395949276,  # Team Principal
        1475933230405259387,  # Vice Team Principal
        1475861814527922258,  # FIA
        1475622175166828554,  # Server Admin
    }

    for role in message.author.roles:
        if role.id in leadership_role_ids:
            return "leadership"

    return "default"


def build_prompt(personality_block, user_message, username):
    return f"""
{personality_block}

The user speaking is: {username}
The user's message is: "{user_message}"

Respond in character. Keep responses short (1–2 sentences). Never break character.
"""



async def generate_reply(user_id, prompt, personality_key):
    if user_id not in USER_SESSIONS:
        USER_SESSIONS[user_id] = []

    # Ensure system prompt is always correct
    if not USER_SESSIONS[user_id] or USER_SESSIONS[user_id][0]["role"] != "system":
        USER_SESSIONS[user_id].insert(0, {"role": "system", "content": PERSONALITY[personality_key]})
    else:
        USER_SESSIONS[user_id][0]["content"] = PERSONALITY[personality_key]

    messages = USER_SESSIONS[user_id]
    messages.append({"role": "user", "content": prompt})

    try:
        response = groq_client.chat.completions.create(
            model=model,
            messages=messages,
            stream=False
        )

        if response.choices and response.choices[0].message:
            reply = response.choices[0].message.content
        else:
            reply = "I couldn't generate a reply."

    except Exception as e:
        reply = f"Error: {e}"

    messages.append({"role": "assistant", "content": reply})
    return reply

@bot.event
async def on_message(message):
    if message.author.bot:
        return

    # Only respond when the bot is mentioned
    if bot.user not in message.mentions:
        return

    # Determine personality
    personality_key = get_personality_for_user(message)

    # Remove the mention tag from the prompt
    clean_prompt = message.content.replace(f"<@{bot.user.id}>", "").strip()

    # Detect if the user is asking for stats/profile
    lowered = clean_prompt.lower()
    wants_profile = any(keyword in lowered for keyword in ["profile", "stats", "info", "data"])

    if wants_profile:
        # Extract a name-like token from the message
        tokens = clean_prompt.replace("'s", "").split()
        candidate = tokens[-1]

        matched_name, profile = get_driver_profile_fuzzy(candidate)

        if matched_name and profile:
            clean_prompt = f"""
The user asked for driver information.

Matched driver: {matched_name}

Driver profile data:
{profile}

Respond using your assigned personality tone.
"""
        else:
            clean_prompt = f"""
The user asked for driver information, but no matching driver was found.
User query: {clean_prompt}

Respond using your assigned personality tone.
"""

    # Generate the AI reply
    reply = await generate_reply(message.author.id, clean_prompt, personality_key)

    await message.channel.send(reply)


@bot.tree.command(
    name="licensereq",
    description="Request a license for an EA ID"
)
@app_commands.describe(
    ea_id="The EA ID to check",
    user="The Discord user who should receive the license"
)
async def licensereq(interaction: discord.Interaction, ea_id: str, user: discord.Member):
    await interaction.response.defer()
    msg = await process_license_request(interaction, ea_id, user)
    await interaction.followup.send(msg)


# -----------------------------
# STARTUP AUTO-PROCESSOR
# -----------------------------

@bot.event
async def on_ready():
    await bot.tree.sync()
    print("Commands synced.")
    print(f"Bot online as {bot.user}")

    # Load caches (async)
    await load_all_caches()

    # Sync slash commands
    try:
        synced = await bot.tree.sync()
        print(f"Slash commands synced: {len(synced)}")
    except Exception as e:
        print(f"Slash sync error: {e}")

    request_channel = bot.get_channel(REQUEST_CHANNEL_ID)

    # Optional: still keep your generic backfill if you want
    # await backfill_message_ids(request_channel)

    # --- STARTUP PROCESSING OF UNPROCESSED REQUESTS ---
    print("Startup processing of unprocessed requests...")

    # Get the settings sheet
    settings_sheet = sheet.spreadsheet.worksheet("settings")

    # Column Z = list of unprocessed Request IDs
    # (adjust index if your Z is not 26 in gspread's 1-based indexing)
    unprocessed_ids = settings_sheet.col_values(26)  # column Z

    # Strip header / blanks
    unprocessed_ids = [x.strip() for x in unprocessed_ids if x.strip().isdigit()]

    for rid_str in unprocessed_ids:
        request_id = int(rid_str)

        # Find the row in the main form sheet
        row = find_row_by_request_id(request_id)
        if row is None:
            print(f"Request {request_id}: no matching row, skipping.")
            continue

        # Check processed flag (col 20)
        processed_value = sheet.cell(row, 20).value or ""
        if processed_value.strip().lower() == "yes":
            print(f"Request {request_id}: already processed, skipping.")
            continue

        # Get message ID (col 21)
        message_id_value = sheet.cell(row, 21).value

        # If no message ID, try to find and log it from channel history
        msg = None
        if not message_id_value:
            print(f"No message ID for request {request_id}, searching history...")
            async for m in request_channel.history(limit=1000):
                mid = extract_request_id(m.content)
                if mid == request_id:
                    msg = m
                    sheet.update_cell(row, 21, str(m.id))
                    print(f"Logged message ID {m.id} for request {request_id}.")
                    break

            if msg is None:
                print(f"Could not find message for request {request_id}, skipping.")
                continue
        else:
            try:
                msg = await request_channel.fetch_message(int(message_id_value))
            except Exception as e:
                print(f"Failed to fetch message {message_id_value} for request {request_id}: {e}")
                continue

        # At this point we have a message and its ID is logged in the sheet

        # Check reactions for approval / rejection
        approved = False
        rejected = False
        approver_member = None

        for reaction in msg.reactions:
            # APPROVAL: ✅ by approver
            if str(reaction.emoji) == "✅":
                users = [u async for u in reaction.users()]
                approvers = [u for u in users if u.id in APPROVER_IDS]
                if approvers:
                    approved = True
                    approver_member = approvers[0]
                    break

            # REJECTION: ❌ by approver
            if str(reaction.emoji) == "❌":
                users = [u async for u in reaction.users()]
                if any(u.id in APPROVER_IDS for u in users):
                    rejected = True

        if approved:
            print(f"Request {request_id}: approved at startup, processing row {row}.")
            await process_row(row, request_channel, approver=approver_member)
        elif rejected:
            print(f"Request {request_id}: rejected at startup, marking as No.")
            sheet.update_cell(row, 20, "No")
        else:
            print(f"Request {request_id}: no approver reaction, leaving unprocessed (but message ID logged).")

    print("Startup processing complete.")


# -----------------------------
# CLEAR ROLES COMMAND
# -----------------------------

def is_approver():
    async def predicate(ctx):
        return ctx.author.id in APPROVER_IDS
    return commands.check(predicate)





@bot.command(name="clearroles")
@is_approver()
async def clear_roles(ctx, *role_names):

    if not role_names:
        await ctx.reply("Please specify at least one role.")
        return

    guild = ctx.guild
    roles_to_clear = []

    for name in role_names:
        role = discord.utils.get(guild.roles, name=name)
        if role:
            roles_to_clear.append(role)
        else:
            await ctx.send(f"Role not found: {name}")

    if not roles_to_clear:
        return

    await ctx.reply("Clearing roles…")

    removed_count = 0

    for member in guild.members:
        to_remove = [r for r in roles_to_clear if r in member.roles]
        if to_remove:
            await member.remove_roles(*to_remove)
            removed_count += 1

    await ctx.send(f"Removed roles from {removed_count} members.")
import random


from discord import app_commands

# Cache for club roles
CLUB_ROLES = []

async def load_club_roles():
    global CLUB_ROLES
    settings_sheet = sheet.spreadsheet.worksheet("settings")
    CLUB_ROLES = settings_sheet.col_values(23)  # Column W is index 23 (1-based)
    CLUB_ROLES = [r.strip() for r in CLUB_ROLES if r.strip()]




# Autocomplete for club names
@bot.tree.command(
    name="clubs",
    description="Join or leave a club by toggling its role."
)
@app_commands.describe(club="Choose a club to join or leave")
async def clubs(interaction: discord.Interaction, club: str):

    await interaction.response.defer(ephemeral=False)

    guild = interaction.guild
    member = interaction.user

    # Ensure cache is loaded
    if not CLUB_ROLES:
        await load_club_roles()

    # Validate club
    if club not in CLUB_ROLES:
        await interaction.followup.send("❌ That club does not exist.")
        return

    # Find the role in Discord
    role = discord.utils.get(guild.roles, name=club)
    if role is None:
        await interaction.followup.send(f"❌ The role **{club}** does not exist in Discord.")
        return

    # Toggle logic
    if role in member.roles:
        try:
            await member.remove_roles(role)
            await interaction.followup.send(f"➖ <@{member.id}> removed from **{club}**")
        except discord.Forbidden:
            await interaction.followup.send("❌ I do not have permission to remove that role.")
    else:
        try:
            await member.add_roles(role)
            await interaction.followup.send(f"➕ <@{member.id}> added to **{club}**")
        except discord.Forbidden:
            await interaction.followup.send("❌ I do not have permission to add that role.")


# Autocomplete handler
@clubs.autocomplete("club")
async def clubs_autocomplete(interaction: discord.Interaction, current: str):
    if not CLUB_ROLES:
        await load_club_roles()

    current = current.lower()
    suggestions = [
        app_commands.Choice(name=club, value=club)
        for club in CLUB_ROLES
        if current in club.lower()
    ]

    return suggestions[:25]


import random


# -----------------------------
# START BOT
# -----------------------------

bot.run(TOKEN)
